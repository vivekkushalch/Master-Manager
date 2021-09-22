# MASTER-MANAGER MADE BY VIVEK KUSHAL CHAKRABORTY

# IMPORTING REQUIRED LIBERIES
from tkinter import *
from tkinter import filedialog, messagebox, colorchooser, font
from tkinter.ttk import Combobox, Progressbar, Style, Treeview
from PIL import ImageTk, Image
import openpyxl as pyxl
import os
import pickle
import datetime
import calendar
import shutil
import threading
import webbrowser




####------- DATABASE INSTANCE LIST --------#####
data = []
####---------------------------------------#####


# temp list to hold the temp data
emp_details = []



# CHECKS THE EXISTANCE OF THE CORE FILES AND ACTS ACCORDINGLY
def file_check():
    # check the existance of DATABASE.xlsx file
    if os.path.exists('DATABASE.xlsx'):
        print(" Foound DATABASE.xlsx at: ", os.path.abspath('DATABASE.xlsx'))
    else:
        print("\n############# Creating New DATABASE.xlsx")
        # creating a new xlsx workbook instance
        wb = pyxl.Workbook()
        wb.properties.creator = "MASTER-MANAGER"
        emp_record = wb.create_sheet('EMP_RECORD', 0)
        sheet2 = wb.create_sheet('BACKUP_EMP_RECORD', 1)
        # saving the database.xlsx workbook
        wb.save("DATABASE.xlsx")

    # check the keys.pkl file
    if os.path.exists("Keys.pkl"):
        print(" Foound Keys.pkl at: ", os.path.abspath('Keys.pkl'))
    else:
        print("####### Creating New KEYS.pkl file")
        # create a new file
        target_file = open('keys.pkl', 'wb')
        # default data
        db = {'admin': '123456'}
        # loading the keys file in pickle
        pickle.dump(db, target_file)
        target_file.close()
        print("saved at: ", os.path.abspath("keys.pkl"))



# EXTARCTS DATA FROM DATABASE.xlsx FILE AND STORES IT IN DATA LIST
def read_xlsx():
    if os.path.exists('DATABASE.xlsx'):
        print(" Foound DATABASE.xlsx at: ", os.path.abspath('DATABASE.xlsx'))
        # loading the database workbook
        wb = pyxl.load_workbook("DATABASE.xlsx")
        # setting setting sheet as the active workbook
        sheet = wb.active
        # storing the row data from the database.xlsx file in data list
        for row in sheet.values:
            data.append(row)
        # removing the headings list    
        data.pop(0)
        # closing the workbook
        wb.close()
    else:
        file_check()



# AUTHINTICATES THE USER UPON LOGIN
def auth(u, p):
    # create new keys.pkl file if it does not exists
    if os.path.exists('keys.pkl'):
        print(" Foound Keys.pkl at: ", os.path.abspath('Keys.pkl'))
        target_file = open('keys.pkl', 'rb')
        db = pickle.load(target_file)
        # authintication logic
        if u in db:
            if p in db[u]:
                print("\nUser found!")
                return 'yes'
        else:
            return 'no'

    else:
        file_check()












# 1st WINDOW/SCREEN FOR SPLASH/LOADING PROGRESS
def splash_screen():
    try:
        # loading data from exel database file in different thread
        read_xlsx_thread = threading.Thread(target=read_xlsx)
        read_xlsx_thread.start()
        # splash screen/window
        splash = Tk()
        # pos and size of splash scereen
        screen_width = splash.winfo_screenwidth()
        screen_height = splash.winfo_screenheight()
        splash.geometry("500x300+"+str(screen_width//2-250) +
                        "+"+str(screen_height//2-150))
        # adding splash image
        img = Image.open("./res/img/logo.jpg")
        img = img.resize((200, 200), Image.ANTIALIAS)
        thumbnail = ImageTk.PhotoImage(img)
        img2 = Image.open("./res/img/login_background.jpg")
        thumbnail2 = ImageTk.PhotoImage(img2)
        # creating canvas for background image
        bg_canvas = Canvas(splash, width=500, height=300)
        bg_canvas.pack(fill='both', expand=True)
        bg_canvas.create_image(0, 0, image=thumbnail2, anchor='nw')
        bg_canvas.create_image(250, 130, image=thumbnail)
        bg_canvas.create_text(250, 250, text="Master-Manager",
                              font="Courier 15", fill="white")
        # overriding to fullscreen
        splash.overrideredirect(True)
        # loding progress bar
        sp_bar = Progressbar(bg_canvas, mode='determinate', length=500)
        sp_bar.pack(side="bottom")
        sp_bar.start(15)
        # stop the progressbar and call login screen 
        def destroy_screen():
            sp_bar.stop()
            login_screen(splash)
        # after aprox 1 sec call destroy_screen method    
        splash.after(1000, destroy_screen)

        splash.mainloop()
    except Exception as ed:
        print("SPLASH SCREEN ERROR: ", ed)




# 2nd SCREEN/WINDOW FOR LOGIN
def login_screen(splash):
    try:
        # destroy the splash screen
        splash.destroy()
    except Exception:
        pass
    # login variable as a tkinter instance
    login = Tk()
    login.title("Master-Manager")
    # window icon
    icon = PhotoImage(file="./res/img/logo.jpg")
    login.iconphoto(False, icon)
    # pos and size of login scereen
    screen_width = login.winfo_screenwidth()
    screen_height = login.winfo_screenheight()
    login.geometry("600x500+"+str(screen_width//2-300) +
                   "+"+str(screen_height//2-300))
    # loading and resizing login bacground image
    img = Image.open("./res/img/login_background.jpg")
    thumbnail = ImageTk.PhotoImage(img)
    # login background canvas with image
    bg_canvas = Canvas(login, width=600, height=500)
    bg_canvas.pack(fill='both', expand=True)
    bg_canvas.create_image(0, 0, image=thumbnail, anchor='nw')
    bg_canvas.create_text(300, 50, text="Login",
                          font="Courier 30", fill="white")
    bg_canvas.create_text(205, 175, text="Username",
                          font="Courier 15", fill="white")
    bg_canvas.create_text(205, 255, text="Password",
                          font="Courier 15", fill="white")
    # Username and Password ENTRY BOX
    username = Entry(login, width=25, font='Courier 15')
    username.place(x=160, y=190)

    pwd = Entry(login, width=25, font='Courier 15', show="*")
    pwd.place(x=160, y=270)
    # show/hide password method 
    def toggle_pwd():
        if pwd.cget('show') == "*":
            pwd.config(show="")
            show_btn.config(text="T")
        else:
            pwd.config(show="*")
            show_btn.config(text="❉")  # ❉
    # Show / hide pwd btn
    show_btn = Button(login, text='❉', width=5,
                      font="Courier 8", command=toggle_pwd)
    show_btn.place(x=419, y=271)

    # Authincating user
    def call_auth():
        u = username.get()
        p = pwd.get()
        res = auth(u, p)
        if res == 'yes':
            # calling mainscreen 
            main_screen(login)
        else:
            messagebox.showerror("Invalid Username and Pasword",
                                 "Please enter a valid Username and Password")

    # Login Button
    login_btn = Button(login, text="Login", font="Courier 12", bg="#982E3C",
                       fg="white", activebackground="#BF1832", command=call_auth, width=10)
    login_btn.place(x=255, y=350)
    # Binding auth func to activate when enter is pressed
    login.bind('<Return>', lambda event=None: login_btn.invoke())

    login.mainloop()





# 3rd SCREEN/WINDOW MAIN SCREEN OF THE APPLICATION  
def main_screen(login):
    # destroying login screen
    login.destroy()

    # main var as a tkinter instance
    main = Tk()
    main.title("MASTER-MANAGER")
    # window icon
    icon = PhotoImage(file="./res/img/logo.jpg")
    main.iconphoto(False, icon)



    ##--- MAIN WINDOW CANVAS --##
    # main screen canvas
    # loading and resizing login bacground image
    img = Image.open("./res/img/login_background.jpg")
    thumbnail = ImageTk.PhotoImage(img)
    # login background image and text
    bg_canvas = Canvas(main, width=1000, height=720)
    bg_canvas.pack(fill='both', expand=True)
    bg_canvas.create_image(0, 0, image=thumbnail, anchor='nw')



    ##----Treeview FORMATING STYLE,FRAME,SCROLLBAR--##
    # styling the treeview (data viewer)
    style = Style()
    style.theme_use('default')
    # setting custom colours to the treeview
    style.configure("Treeview", background="#D3D3D3",
                    foreground="black", rowheight=28, fieldbackground="#D3D3D3")
    # setting the custom selection colour of treeview
    style.map("Treeview", background=[("selected", "#c22525")])
    # treeview frame
    tv_frame = Frame(bg_canvas)
    tv_frame.pack(pady=10)
    # treeview scrollbar
    tv_scroll_bar = Scrollbar(tv_frame)
    tv_scroll_bar.pack(side=RIGHT, fill=Y)
    # Treeview
    tv = Treeview(tv_frame, yscrollcommand=tv_scroll_bar.set,
                  selectmode="extended")
    tv.pack()
    # setting command for tv_scroll_bar
    tv_scroll_bar.config(command=tv.yview)

    # treeview columns
    tv['columns'] = ("SL NO", "ID", 'FIRST NAME', 'LAST NAME',
                     'DEPARTMENT', 'TEAM', 'POST', 'SHIFT', 'PAY')
    # FORMATING TREEVIEW COLUMNS
    tv.column('#0', width=0, stretch=NO)
    tv.column("SL NO", anchor=CENTER, width=100)
    tv.column("ID", anchor=CENTER, width=150)
    tv.column("FIRST NAME", anchor='w', width=150)
    tv.column("LAST NAME", anchor='w', width=150)
    tv.column("DEPARTMENT", anchor=CENTER, width=150)
    tv.column("TEAM", anchor=CENTER, width=150)
    tv.column("POST", anchor=CENTER, width=150)
    tv.column("SHIFT", anchor=CENTER, width=150)
    tv.column("PAY", anchor=CENTER, width=150)
    # Treeview column headings
    tv.heading("#0", text="", anchor='w')
    tv.heading("SL NO", text="SL NO", anchor=CENTER)
    tv.heading("ID", text="ID", anchor=CENTER)
    tv.heading("FIRST NAME", text="FIRST NAME", anchor='w')
    tv.heading("LAST NAME", text="LAST NAME", anchor='w')
    tv.heading("DEPARTMENT", text="DEPARTMENT", anchor=CENTER)
    tv.heading("TEAM", text="TEAM", anchor=CENTER)
    tv.heading("POST", text="POST", anchor=CENTER)
    tv.heading("SHIFT", text="SHIFT", anchor=CENTER)
    tv.heading("PAY", text="PAY", anchor=CENTER)




    ##---- TREEVIEW ---##
    # treeview stryped rows
    tv.tag_configure('oddrow', background="white")
    tv.tag_configure('evenrow', background="#fa8c8c")



    # adding data to treeview from database variable data
    def refresh_treeview():
        # delete all elements in treeview
        for rows in tv.get_children():
            tv.delete(rows)
        # adding data to treeview
        count = 0
        for record in data:
            if count % 2:
                # record[] = data index
                tv.insert(parent='', index='end', iid=count, text='', values=(
                    str(count+1), record[1], record[2], record[3], record[4], record[5], record[6], record[7], record[8]), tags=('evenrow',))
            else:
                tv.insert(parent='', index='end', iid=count, text='', values=(
                    str(count+1), record[1], record[2], record[3], record[4], record[5], record[6], record[7], record[8]), tags=('oddrow',))
            count += 1
    refresh_treeview()



    ##---- LABEL FRAME FOR LABELS,ENTRYBOX AND BUTTONS ----##

    # Employee Records edit entry Frame and entry boxes
    rec_frame = LabelFrame(main, text="Employee Details")
    # bg_canvas.create_window(500,350,window=rec_frame)
    rec_frame.pack(fill=X, expand='yes', padx=20)



    ##---- LABELS AND ENTRY BOX FOR DATA MANUPULATION-----##
    id_label = Label(rec_frame, text='ID', font="Courier 12")
    id_label.grid(row=1, column=0, padx=10, pady=10)
    id_entry = Entry(rec_frame, font="Courier 12", width=18)
    id_entry.grid(row=1, column=1, padx=10, pady=10)

    first_name_label = Label(rec_frame, text='First Name', font="Courier 12")
    first_name_label.grid(row=0, column=0, padx=10, pady=10)
    first_name_entry = Entry(rec_frame, font="Courier 12", width=18)
    first_name_entry.grid(row=0, column=1, padx=10, pady=10)

    last_name_label = Label(rec_frame, text='Last Name', font="Courier 12")
    last_name_label.grid(row=0, column=2, padx=10, pady=10)
    last_name_entry = Entry(rec_frame, font="Courier 12", width=18)
    last_name_entry.grid(row=0, column=3, padx=10, pady=10)

    department_label = Label(rec_frame, text='Department', font="Courier 12")
    department_label.grid(row=1, column=2, padx=10, pady=10)
    department_entry = Entry(rec_frame, font="Courier 12", width=18)
    department_entry.grid(row=1, column=3, padx=10, pady=10)

    team_label = Label(rec_frame, text='Team', font="Courier 12")
    team_label.grid(row=2, column=0, padx=10, pady=10)
    team_entry = Entry(rec_frame, font="Courier 12", width=18)
    team_entry.grid(row=2, column=1, padx=10, pady=10)

    post_label = Label(rec_frame, text='Post', font="Courier 12")
    post_label.grid(row=2, column=2, padx=10, pady=10)
    post_entry = Entry(rec_frame, font="Courier 12", width=18)
    post_entry.grid(row=2, column=3, padx=10, pady=10)

    shift_label = Label(rec_frame, text='Shift', font="Courier 12")
    shift_label.grid(row=3, column=0, padx=10, pady=10)
    shift_entry = Entry(rec_frame, font="Courier 12", width=18)
    shift_entry.grid(row=3, column=1, padx=10, pady=10)

    pay_label = Label(rec_frame, text='Pay', font="Courier 12")
    pay_label.grid(row=3, column=2, padx=10, pady=10)
    pay_entry = Entry(rec_frame, font="Courier 12", width=18)
    pay_entry.grid(row=3, column=3, padx=10, pady=10)

    search_entry = Entry(rec_frame, font="Courier 12", width=15)
    search_entry.insert(0,'Search')
    search_entry.grid(row=3, column=8, padx=10, pady=10)

    short_cut_lb = Label(rec_frame, text='Keyboard Shortcuts', font="Courier 12")
    short_cut_lb.grid(row=0, column=10, padx=10, pady=10)

    short_cut_move_up_lb = Label(rec_frame, text='Left Arrow Move Up', font="Courier 12")
    short_cut_move_up_lb.grid(row=1, column=10, padx=10, pady=10)

    short_cut_move_down_lb = Label(rec_frame, text='Right Arrow Move Down', font="Courier 12")
    short_cut_move_down_lb.grid(row=2, column=10, padx=10, pady=10)

    short_cut_tab_lb = Label(rec_frame, text='TAB Select Entry box', font="Courier 12")
    short_cut_tab_lb.grid(row=3, column=10, padx=10, pady=10)



    ##------ FUNCTIONS/METHODS -------##


    # updating the data in database.xlsx file
    def update_db_file():
        # loading database.xlsx file
        wb = pyxl.load_workbook('DATABASE.xlsx')
        # getting the sheetnames   ['EMP_RECORD', 'BACKUP_EMP_RECORD', 'Sheet']
        shts = wb.sheetnames
        # Getting Sheet by name
        emp_ws = wb['EMP_RECORD']
        # activating EMP sheet
        wb.active = emp_ws
        emp_sheet = wb.active
        # delete all data
        emp_sheet.delete_rows(2, int(emp_sheet.max_row))

        # writing new data
        for record in data:
            emp_sheet.append(record)
            # print(record)

        # saving the file
        wb.save('DATABASE.xlsx')

    # updates the data list with live data from the treeview widget
    def update_live_data_to_data_list():
        new_data = []
        for r in tv.get_children():
            new_data.append(tv.item(r)['values'])
        data.clear()
        for i in new_data:
            data.append(i)


    # updating data in Tkinter.Treeview widget
    def update_data():
        # grabbing entry data
        emp_data = []
        ids = id_entry.get()
        fn = first_name_entry.get()
        ln = last_name_entry.get()
        dep = department_entry.get()
        tm = team_entry.get()
        pst = post_entry.get()
        sft = shift_entry.get()
        pd = pay_entry.get()

        if ids == '':
            messagebox.showerror("Master-Manager", 'ID Entry is empty')
        elif fn == '':
            messagebox.showerror("Master-Manager", 'FIRST NAME Entry is empty')
        elif ln == '':
            messagebox.showerror("Master-Manager", 'LAST NAME Entry is empty')
        elif dep == '':
            messagebox.showerror("Master-Manager", 'DEPARTMENT Entry is empty')
        elif tm == '':
            messagebox.showerror("Master-Manager", 'TEAM Entry is empty')
        elif pst == '':
            messagebox.showerror("Master-Manager", 'POST Entry is empty')
        elif sft == '':
            messagebox.showerror("Master-Manager", 'SHIFT Entry is empty')
        elif pd == '':
            messagebox.showerror("Master-Manager", 'PAY Entry is empty')
        else:
            # updating row data in treeview widget
            selected = tv.focus()
            tv.item(selected, text='', values=(id_entry.get(), first_name_entry.get(), last_name_entry.get(
            ), department_entry.get(), team_entry.get(), post_entry.get(), shift_entry.get(), pay_entry.get()))
            emp_data = ['',ids, fn, ln, dep, tm, pst, sft, pd]
            # updating new data in database variable data
            data[int(selected)] = emp_data
            # refreshing treeview widget
            refresh_treeview()
            # clearing entry boxes
            id_entry.delete(0, END)
            first_name_entry.delete(0, END)
            last_name_entry.delete(0, END)
            department_entry.delete(0, END)
            team_entry.delete(0, END)
            post_entry.delete(0, END)
            shift_entry.delete(0, END)
            pay_entry.delete(0, END)

            # updating data in DATABASE.xlsx file 
            update_db_file()


    # remove rows in treeview widget
    def delete_rows():
        # grabing the row number
        selected = tv.focus()
        # grabing all the details of the selected employee
        employee_details = tv.item(selected, 'values')
        # grabbing the target row number
        target = tv.selection()
        # setting default value for response variable res
        res = ''
        if len(target) == 1:
            # asking for conformation
            res = messagebox.askquestion("Master-Manager", "Are you sure?\nRemove "+'SL_NO: '+str(employee_details[0])+" "+str(employee_details[1])+" "+str(
                employee_details[2])+" "+str(employee_details[3])+" "+str(employee_details[4])+" "+str(employee_details[5])+" "+str(employee_details[6]))
        if len(target) > 1:
            res = messagebox.askquestion(
                "Master-Manager", "Are you sure?\nRemove "+str(len(target))+" Employees")
        else:
            pass
        if res == 'yes':
            # deleting rows in treeview widget
            for i in target:
                try:
                    # deleting items form treeview widget
                    tv.delete(i)
                    # logic to repopulate data list with updated data
                    update_live_data_to_data_list()

                except Exception as ex:
                    print("DELETE MULTIPLE ROWS ERROR line 489: ",ex)    

            # refreshing treeview widget
            refresh_treeview()

            # clearing entry boxes
            id_entry.delete(0, END)
            first_name_entry.delete(0, END)
            last_name_entry.delete(0, END)
            department_entry.delete(0, END)
            team_entry.delete(0, END)
            post_entry.delete(0, END)
            shift_entry.delete(0, END)
            pay_entry.delete(0, END)
        else:
            pass

    # delete all data from treeview widget
    def del_all_data():
        # method to del all data only after authintication
        def delete_database_auth_done():
            res = messagebox.askquestion(
                "Master-Manager", 'Are you sure?\nDeleted data cannot be recovered later !')
            if res == 'yes':
                for rows in tv.get_children():
                    tv.delete(rows)
                data.clear()
            else:
                pass

        # user Authintication for Data deletion        
        def auth_for_del_database():
            # tkinter toplevel window instance 
            auth_del = Toplevel(main)
            auth_del.geometry('400x200')
            auth_del.title("Authintication")
            auth_del.config(bg='#d94856')
            auth_del.iconphoto(False, icon)
            # Labels
            un = Label(auth_del, text='Username',
                       font="Courier 11", bg='#d94856', fg='white')
            un.grid(row=0, column=0, padx=10, pady=10)

            pas = Label(auth_del, text='Password',
                        font="Courier 11", bg='#d94856', fg='white')
            pas.grid(row=1, column=0, padx=10, pady=10)
            # Entry box
            user_name = Entry(auth_del, font="Courier 11", width=27)
            user_name.grid(row=0, column=1, padx=10, pady=10)

            pwd_entry = Entry(auth_del, font="Courier 11", width=27, show="*")
            pwd_entry.grid(row=1, column=1, padx=10, pady=10)

            # cross-checking Username and Password from keys.pkl file
            def check_uid():
                u_name = user_name.get()
                pws = pwd_entry.get()
                # loading keys.pkl file
                target_file = open('keys.pkl', 'rb')
                auth_db = pickle.load(target_file)
                # auth logic
                if u_name in auth_db:
                    if pws in auth_db[u_name]:
                        print("\n\nVALID USER FOR DELETE DATABASE\n\n")
                        auth_del.destroy()
                        delete_database_auth_done()

                else:
                    auth_del.destroy()
                    messagebox.showerror(
                        'MASTER-MANAGER', 'Invalid Username or Password')
            # button
            auth_btn = Button(auth_del, text='Login', font="Courier 11",
                              bg='#d94856', fg='white', command=check_uid)
            auth_btn.grid(row=4, column=1, padx=10, pady=10)

            auth_del.mainloop()
        warning = messagebox.showwarning(
            "Master-Manager", 'All the Employee data will be deleted !')
        if warning:
            auth_for_del_database()
        else:
            pass


    # clears all the data from all entry boxe widgets    
    def clear_entry_box():
        id_entry.delete(0, END)
        first_name_entry.delete(0, END)
        last_name_entry.delete(0, END)
        department_entry.delete(0, END)
        team_entry.delete(0, END)
        post_entry.delete(0, END)
        shift_entry.delete(0, END)
        pay_entry.delete(0, END)
        search_entry.delete(0, END)


    # adds new data to the DATABASE.xlsx file
    def add_new_emp():
        # new list for grabbing all data fields
        emp_data = []
        # grabbing all data from all data fields
        ids = id_entry.get()
        fn = first_name_entry.get()
        ln = last_name_entry.get()
        dep = department_entry.get()
        tm = team_entry.get()
        pst = post_entry.get()
        sft = shift_entry.get()
        pd = pay_entry.get()
        # Error protection logic
        if ids == '':
            messagebox.showerror("Master-Manager", 'ID Entry is empty')
        elif fn == '':
            messagebox.showerror("Master-Manager", 'FIRST NAME Entry is empty')
        elif ln == '':
            messagebox.showerror("Master-Manager", 'LAST NAME Entry is empty')
        elif dep == '':
            messagebox.showerror("Master-Manager", 'DEPARTMENT Entry is empty')
        elif tm == '':
            messagebox.showerror("Master-Manager", 'TEAM Entry is empty')
        elif pst == '':
            messagebox.showerror("Master-Manager", 'POST Entry is empty')
        elif sft == '':
            messagebox.showerror("Master-Manager", 'SHIFT Entry is empty')
        elif pd == '':
            messagebox.showerror("Master-Manager", 'PAY Entry is empty')
        else:
            # storing all fields data into a temp list
            emp_data = [str(len(data)+1), ids, fn, ln, dep, tm, pst, sft, pd]
            # appending new data to the database instance variable data
            data.append(emp_data)
            # refreshing treeview widget
            refresh_treeview()
            # updating DATABASE.xlsx file
            update_db_file()


    # method to fill the fields Entry boxes when clicked on a treeview row
    def select_employee_data(event):
        # clearing entry boxes
        id_entry.delete(0, END)
        first_name_entry.delete(0, END)
        last_name_entry.delete(0, END)
        department_entry.delete(0, END)
        team_entry.delete(0, END)
        post_entry.delete(0, END)
        shift_entry.delete(0, END)
        pay_entry.delete(0, END)
        # grabing the employee number
        selected = tv.focus()
        # grabing all the details of the selected employee
        employee_details = tv.item(selected, 'values')
        # exception may occure if there is more than one emp record selected
        try:
            # filling the entry boxes with the grabbed data
            id_entry.insert(0, employee_details[1])
            first_name_entry.insert(0, employee_details[2])
            last_name_entry.insert(0, employee_details[3])
            department_entry.insert(0, employee_details[4])
            team_entry.insert(0, employee_details[5])
            post_entry.insert(0, employee_details[6])
            shift_entry.insert(0, employee_details[7])
            pay_entry.insert(0, employee_details[8])
        except Exception:
            pass
            clear_entry_box()   


    # temp list to store moved item index for moving row up
    data_index = []
    # temp list for moved index counting for moving row up
    move_ct = []

    # method to move row up in treeview widget
    def move_row_up():
        # grabbing selected row
        rows = tv.selection()
        # moving the row up in treeview widget
        for row in rows:
            tv.move(row, tv.parent(row), tv.index(row)-1)
            # temp var to store row val
            idx = row
            # appending moved item index
            data_index.append(data.index(data[int(row)]))
            # incrementing the moved to index count 
            move_ct.append(1)
        # logic for moving item in list for old to new index    
        if int(data_index[0])-len(move_ct) < 0:
            print("INDEX OF ITEM MOVED: ", data_index[0], "NEW INDEX: ", 0)
        else:
            print("INDEX OF ITEM MOVED: ", data_index[0], "NEW INDEX: ", int(
                data_index[0])-len(move_ct))
        
        # logic to move the treeview up without loosing focus/selection
        update_live_data_to_data_list()
        
        
    # move row down
    def move_row_down():
        # selected row
        rows = tv.selection()
        # moving the row down in treeview widget
        for row in reversed(rows):
            tv.move(row, tv.parent(row), tv.index(row)+1)

        # logic to move the treeview down without loosing focus/selection  
        update_live_data_to_data_list()




    # Dynamic search method
    def search_emp():
        search_result = []
        find = str(search_entry.get())
        if find == '':
            refresh_treeview()
        if find != '':
            for emp_record in data:
                for item in emp_record:
                    if find in str(item):
                        search_result.append(emp_record)
                    else:
                        continue
            ## updating the treeview with search result ##

            # delete all elements in treeview
            for rows in tv.get_children():
                tv.delete(rows)
            # adding data to treeview
            count = 0
            for record in search_result:
                # print(data.index(record))
                if count % 2:
                    # record[] = column index
                    tv.insert(parent='', index='end', iid=count, text='', values=(
                        record[0], record[1], record[2], record[3], record[4], record[5], record[6], record[7], record[8]), tags=('evenrow',))
                else:
                    tv.insert(parent='', index='end', iid=count, text='', values=(
                        record[0], record[1], record[2], record[3], record[4], record[5], record[6], record[7], record[8]), tags=('oddrow',))
                count += 1
            if len(search_result) == 0:
                messagebox.showerror('NOT FOUND', str(find)+' not found')
                refresh_treeview()
            else:
                messagebox.showinfo('Search', str(len(search_result))+' Found')
        else:
            refresh_treeview()

    # Binding treeview with mouse left click to call the select_employee data func
    tv.bind('<ButtonRelease-1>', select_employee_data)


    # method to destroy main screen and run login screen
    def exit_to_login():
        main.destroy()
        login_screen('none')

    # show sats of the database and application
    def database_info():
        def db_info_dashboard():
            info = Toplevel(main)
            info.geometry("550x450")
            info.iconphoto(False, icon)
            info.config(bg='#303030')
            info.title("MASTER-MANAGER/DASHBOARD")
            # db info
            db_info_frame = LabelFrame(
                info, text='DATABASE.xlsx', font="Courier 12", bg='#303030', fg='white')
            db_info_frame.pack(fill='both', expand='yes', padx=20, pady=20)
            # label entry and button
            total_emp = Label(db_info_frame, text='TOTAL         :',
                              font="Courier 12", bg='#303030', fg='white')
            total_emp.grid(row=0, column=0, padx=10, pady=10)

            t_emp_data = Label(db_info_frame, text=str(
                len(data))+' EMPLOYEES', font="Courier 12", bg='#303030', fg='white')
            t_emp_data.grid(row=0, column=1, padx=10, pady=10)

            save_path_label = Label(
                db_info_frame, text='SAVE PATH     :', font="Courier 12", bg='#303030', fg='white')
            save_path_label.grid(row=1, column=0, padx=10, pady=10)

            save_path_entry = Entry(db_info_frame, font="Courier 11", width=27)
            save_path_entry.insert(0, str(os.path.abspath('DATABASE.xlsx')))
            save_path_entry.grid(row=1, column=1, padx=10, pady=10)

            # find  the day
            def findDay(date):
                born = datetime.datetime.strptime(date, '%d/%m/%Y').weekday()
                return (calendar.day_name[born])

            # loading the database workbook
            wb = pyxl.load_workbook("DATABASE.xlsx")
            # Getting Sheet by name
            emp_ws = wb['EMP_RECORD']
            # activating EMP sheet
            wb.active = emp_ws
            emp_sheet = wb.active
            # max row and column
            max_row = emp_sheet.max_row
            max_col = emp_sheet.max_column
            # date created
            dt_created = wb.properties.created.strftime('%d/%m/%Y')
            day_created = findDay(dt_created)
            # last modified
            dt_modified = wb.properties.modified.strftime('%d/%m/%Y')
            day_modified = findDay(dt_modified)
            
            # dimentions of the database
            dimention_label = Label(
                db_info_frame, text='DIMENTIONS    :', font="Courier 12", bg='#303030', fg='white')
            dimention_label.grid(row=2, column=0, padx=10, pady=10)

            db_dimentions_lb = Label(db_info_frame, text=str(
                max_row)+' Rows & '+str(max_col)+' columns', font="Courier 12", bg='#303030', fg='white')
            db_dimentions_lb.grid(row=2, column=1, padx=10, pady=10)

            # date created LABEL
            date_created_label = Label(
                db_info_frame, text='DATE CREATED  :', font="Courier 12", bg='#303030', fg='white')
            date_created_label.grid(row=3, column=0, padx=10, pady=10)

            DT_label = Label(db_info_frame, text=dt_created+" " +
                             day_created, font="Courier 12", bg='#303030', fg='white')
            DT_label.grid(row=3, column=1, padx=10, pady=10)

            # last modified LABEL
            last_modified_label = Label(
                db_info_frame, text='LAST MODIFIED :', font="Courier 12", bg='#303030', fg='white')
            last_modified_label.grid(row=4, column=0, padx=10, pady=10)

            l_modified = Label(db_info_frame, text=dt_modified+" " +
                               day_modified, font="Courier 12", bg='#303030', fg='white')
            l_modified.grid(row=4, column=1, padx=10, pady=10)

            # loading username and pasword from keys.xlsx file
            tg_file = open('keys.pkl', 'rb')
            db = pickle.load(tg_file)
            db_key = list(db.keys())[0]
            db_val = db[db_key]
            tg_file.close()

            # saving new creditientials
            def save_new_auth():
                db.clear()
                new_key = user_entry.get()
                new_val = pwd_entry.get()
                db[new_key] = new_val
                # opeaning target file
                target_file = open('keys.pkl', 'wb')
                # dumping new data
                pickle.dump(db, target_file)
                target_file.close()
                info.destroy()
                messagebox.showinfo(
                    "MASTER-MANAGER", 'USERNAME AND PASSWORD SAVED')

            # username label
            user_label = Label(db_info_frame, text='USERNAME      :',
                               font="Courier 12", bg='#303030', fg='white')
            user_label.grid(row=5, column=0, padx=10, pady=10)
            # username entrybox
            user_entry = Entry(db_info_frame, width=25, font="Courier 12")
            user_entry.insert(0, db_key)
            user_entry.grid(row=5, column=1, padx=10, pady=10)

            # pasword label 
            pwd_label = Label(db_info_frame, text='PASSWORD      :',
                              font="Courier 12", bg='#303030', fg='white')
            pwd_label.grid(row=6, column=0, padx=10, pady=10)
            # password entrybox
            pwd_entry = Entry(db_info_frame, width=25, font="Courier 12")
            pwd_entry.insert(0, db_val)
            pwd_entry.grid(row=6, column=1, padx=10, pady=10)

            # save username and password button
            save_btn = Button(db_info_frame, text='SAVE', font="Courier 12",
                              width=10, bg='#d94856', fg='white', command=save_new_auth)
            save_btn.grid(row=7, column=1, pady=30)

            info.mainloop()
        
        # authintication screen for dashboard
        def dashboard_auth():
            # toplevel window for auth
            auth_del = Toplevel(main)
            auth_del.geometry('400x200')
            auth_del.title("Authintication")
            auth_del.configure(bg='#d94856')
            auth_del.iconphoto(False, icon)
            # Labels
            un = Label(auth_del, text='Username',
                       font="Courier 11", bg='#d94856', fg='white')
            un.grid(row=0, column=0, padx=10, pady=10)

            pas = Label(auth_del, text='Password',
                        font="Courier 11", bg='#d94856', fg='white')
            pas.grid(row=1, column=0, padx=10, pady=10)
            # Entry box
            user_name = Entry(auth_del, font="Courier 11", width=27)
            user_name.grid(row=0, column=1, padx=10, pady=10)

            pwd_entrys = Entry(auth_del, font="Courier 11", width=27, show="*")
            pwd_entrys.grid(row=1, column=1, padx=10, pady=10)

            # authintication logic
            def check_uid():
                u_name = user_name.get()
                pws = pwd_entrys.get()
                # auth logic
                target_file = open('keys.pkl', 'rb')
                auth_db = pickle.load(target_file)
                if u_name in auth_db:
                    if pws in auth_db[u_name]:
                        print("\n\nVALID USER FOR DB DASHBOARD\n\n")
                        auth_del.destroy()
                        db_info_dashboard()
                else:
                    auth_del.destroy()
                    messagebox.showerror(
                        'MASTER-MANAGER', 'Invalid Username or Password')
            # button
            auth_btn = Button(auth_del, text='Login', font="Courier 11",
                              bg='#d94856', fg='white', command=check_uid)
            auth_btn.grid(row=4, column=1, padx=10, pady=10)
            auth_del.mainloop()
        dashboard_auth()


    # method to export DATABASE.xlsx file in desired location
    def export_db_file():
        if os.path.exists('DATABASE.xlsx'):
            original_path = os.path.abspath('DATABASE.xlsx')
            target_path = filedialog.askdirectory()
            shutil.copyfile(original_path, target_path+'\DATABASE.xlsx')
            messagebox.showinfo(
                'MASTER-MANAGER', 'File Exported to\n'+target_path+'/DATABASE.xlsx')
        else:
            messagebox.showerror('MASTER-MANAGER', 'DATABASE.xlsx NOT FOUND')


    ##------- BUTTONS FOR MAIN SCREEN/WINDOW ------##
    update_btn = Button(rec_frame, text='UPDATE DETAILS', font="Courier 10", width=15, command=update_data).grid(
        row=1, column=8, padx=10, pady=10)
    add_btn = Button(rec_frame, text='ADD NEW EMPLOYEE', font="Courier 10", command=add_new_emp).grid(
        row=0, column=8, padx=10, pady=10)
    del_selected_btn = Button(rec_frame, text='REMOVE EMPLOYEE', font="Courier 10", width=15, command=delete_rows).grid(
        row=0, column=9, padx=10, pady=10)
    clear_btn = Button(rec_frame, text='CLEAR ENTRY', font="Courier 10", width=16,
                       command=clear_entry_box).grid(row=2, column=8, padx=10, pady=10)

    move_up_btn = Button(rec_frame, text='▲', width=15, command=move_row_up)
    move_up_btn.grid(row=1, column=9, padx=11, pady=10)

    move_down_btn = Button(rec_frame, text='▼',
                           width=15, command=move_row_down)
    move_down_btn.grid(row=2, column=9, padx=10, pady=10)

    search_btn = Button(rec_frame, text='Refresh',
                        font="Courier 10", width=15, command=search_emp)
    search_btn.grid(row=3, column=9)


    # Dynamically change To and From Refresh and Search Buttons
    def change_btn(event):
        if search_entry.get():
            search_btn.config(text='SEARCH')
        else:
            search_btn.config(text='Refresh')

    # open view help page
    def open_view_help():
        pdf_path = os.path.abspath("./DOCS/MASTER-MANAGER.pdf")
        try:  
            os.system('start chrome "file:///'+pdf_path+'#page=3"')
        except Exception:
            webbrowser.open_new_tab('file://'+pdf_path+'#page=3')
    # open about page
    def open_about():
        pdf_path = os.path.abspath("./DOCS/MASTER-MANAGER.pdf")
        try:
            os.system('start chrome "file:///'+pdf_path+'#page=2"')
        except Exception:
            webbrowser.open('file://'+pdf_path+'#page=2') 
    # open feedback page        
    def open_feedback():
        try:
            os.system('start chrome "https://forms.gle/HJyzk3By4qB2kJKb9"')        
        except Exception:
            webbrowser.open("https://forms.gle/pgU7UD3qpNFSjbBc8")       

            
                  
    ###----- KEY BINDINGS ------####
    search_entry.bind('<Key>', change_btn)
    search_entry.bind('<Return>', lambda event=None: search_btn.invoke())

    tv.bind('<Left>', lambda event=None: move_up_btn.invoke())
    tv.bind('<Right>', lambda event=None: move_down_btn.invoke())



    # menu tool clip
    menu = Menu(main)
    main.config(menu=menu)
    # file menu
    file_menu = Menu(menu, tearoff=False)
    menu.add_cascade(label='File', menu=file_menu)
    file_menu.add_command(label='Dashboard', command=database_info)
    file_menu.add_command(label='Refresh Database', command=refresh_treeview)
    file_menu.add_command(label='Export Database', command=export_db_file)
    file_menu.add_command(label='Delete Database    ', command=del_all_data)
    file_menu.add_separator()
    file_menu.add_command(label='Exit', command=exit_to_login)
    # help menu
    help_menu = Menu(menu, tearoff=False)
    menu.add_cascade(label='Help', menu=help_menu)
    help_menu.add_command(label='View Help',command=open_view_help)
    help_menu.add_command(label='Send Feedback',command=open_feedback)
    help_menu.add_command(label='About Master-Manager',command=open_about)
    help_menu.add_separator()
    help_menu.add_command(label='Made By Vivek Kushal Chakraborty')

    # pos and size of main screen
    screen_width = main.winfo_screenwidth()
    screen_height = main.winfo_screenheight()
    main.geometry(str(screen_width)+"x"+str(screen_height)+"+0+0")

    # method to save the Changed data to DATABASE.xlsx file upon exit Application
    def close_main_window():
        res = messagebox.askyesnocancel("Master-Manager", 'SAVE CHANGES ?')
        if res:
            update_db_file()
            main.destroy()
        if res == False:
            main.destroy()
        else:
            pass

    # call save data upon exit method upon clicking X button on the top right corner        
    main.protocol("WM_DELETE_WINDOW", close_main_window)

    main.mainloop()



# 1st SCREEN/WINDOW TO START THE APPLICATION
splash_screen()
